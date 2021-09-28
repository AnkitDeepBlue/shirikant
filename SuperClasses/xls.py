import openpyxl
from openpyxl.workbook import workbook


class myexel():
    def maxrow(self, path, SheetName):
        # Loading the xl file
        book = openpyxl.load_workbook(path)

        #Getting sheet
        sheet = book.get_sheet_by_name(SheetName)

        #Getting maxkimum row
        rows = sheet.max_row

        #Returning row
        return rows

    def maxcol(self, path, SheetName):
        # Loading the xl file
        book = openpyxl.load_workbook(path)

        sheet = book.get_sheet_by_name(SheetName)

        col=sheet.max_column

        return col

    def getCellValue(self, path, rownum, columnnum):
        book = openpyxl.load_workbook(path)
        sheet = book.active
        return sheet.cell(row=rownum, column=columnnum).value





